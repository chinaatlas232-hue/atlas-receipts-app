import base64
import datetime
import io
import os
import openpyxl
import pandas as pd
import requests
import streamlit as st

st.set_page_config(page_title="وصل تسليم بضاعة - أطلس", layout="wide")

# --- معرفات الملفات من Google Drive ---
SHIPMENT_FILE_ID = "1IESujqsd6-4RbEfr9cnx8xeYNq-WvTUj"
TEMPLATE_FILE_ID = "1_DxNo3KIWWdSQ-Q4r_hatsYZ0sYT8ier"
LOGO_FILE_ID = "1fAz46COaR6SgT9DNbYqx9Ea9iclQEQTA"
CUSTOMER_INFO_FILE_ID = "1gCjzU7Gx5alpv7KZY1mxjIVDJO-yvzww"

# --- مسارات التخزين المؤقت المحلية ---
UPLOAD_DIR = "saved_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

shipment_path = os.path.join(UPLOAD_DIR, "shipments_data.xlsx")
template_path = os.path.join(UPLOAD_DIR, "template.xlsx")
logo_path = os.path.join(UPLOAD_DIR, "logo.png")
customer_info_path = os.path.join(UPLOAD_DIR, "customer_info.xlsx")


# --- دالة موثوقة لتنزيل الملفات من Google Drive ---
def download_file_from_google_drive(file_id, destination):
  URL = "https://docs.google.com/uc?export=download"
  session = requests.Session()
  response = session.get(URL, params={"id": file_id}, stream=True)

  token = None
  for key, value in response.cookies.items():
    if key.startswith("download_warning"):
      token = value
      break

  if token:
    params = {"id": file_id, "confirm": token}
    response = session.get(URL, params=params, stream=True)

  with open(destination, "wb") as f:
    for chunk in response.iter_content(32768):
      if chunk:
        f.write(chunk)


@st.cache_data(ttl=300)
def download_files_from_drive():
  try:
    download_file_from_google_drive(SHIPMENT_FILE_ID, shipment_path)
    download_file_from_google_drive(TEMPLATE_FILE_ID, template_path)
    download_file_from_google_drive(LOGO_FILE_ID, logo_path)
    download_file_from_google_drive(CUSTOMER_INFO_FILE_ID, customer_info_path)
    return True
  except Exception as e:
    st.error(f"خطأ أثناء السحب من درايف: {e}")
    return False


success_sync = download_files_from_drive()

# --- تنسيق الشريط الجانبي ---
st.markdown(
    """
    <style>
    [data-testid="stSidebar"] {
        background-color: #334155;
        color: #f8fafc;
    }
    [data-testid="stSidebar"] h1, 
    [data-testid="stSidebar"] h2, 
    [data-testid="stSidebar"] h3, 
    [data-testid="stSidebar"] label, 
    [data-testid="stSidebar"] .stMarkdown {
        color: #f8fafc !important;
    }
    [data-testid="stSidebar"] button[kind="secondary"] {
        background-color: #991b1b !important;
        color: white !important;
        border: 1px solid #7f1d1d !important;
        font-weight: bold !important;
    }
    </style>
""",
    unsafe_allow_html=True,
)

with st.sidebar:
  st.markdown(
      """
        <div style="text-align: right; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 1px solid #475569;">
            <h2 style="color: #f8fafc; font-size: 16px; margin: 0;">📦 وصل تسليم البضائع</h2>
            <p style="color: #cbd5e1; font-size: 12px; margin: 3px 0 0 0;">(شركة أطلس المحيط)</p>
        </div>
    """,
      unsafe_allow_html=True,
  )

  st.header("⚙️ إدارة الملفات")
  if success_sync:
    st.success("✅ متصل بـ Google Drive بنجاح")
  else:
    st.error("⚠️ فشل الاتصال، يجدر التحقق من الروابط.")

  if st.button("🔄 تحديث البيانات وسحبها من درايف"):
    st.cache_data.clear()
    download_files_from_drive()
    st.rerun()

  ship_mtime = (
      os.path.getmtime(shipment_path) if os.path.exists(shipment_path) else 0
  )
  cust_mtime = (
      os.path.getmtime(customer_info_path)
      if os.path.exists(customer_info_path)
      else 0
  )


  @st.cache_data(show_spinner=False)
  def load_and_merge_data(s_time, c_time):
    if not os.path.exists(shipment_path):
      return None

    try:
      df_s = pd.read_excel(shipment_path)
    except Exception:
      return None

    df_s.columns = df_s.columns.astype(str).str.strip()

    if os.path.exists(customer_info_path):
      try:
        if customer_info_path.endswith(".csv"):
          df_c = pd.read_csv(customer_info_path)
        else:
          df_c = pd.read_excel(customer_info_path)

        df_c.columns = df_c.columns.astype(str).str.strip()

        ship_code_col = next(
            (c for c in df_s.columns if "كود" in c or "code" in c.lower()),
            "الكود",
        )
        cust_code_col = next(
            (c for c in df_c.columns if "كود" in c or "code" in c.lower()),
            "الكود",
        )

        if ship_code_col in df_s.columns and cust_code_col in df_c.columns:
          df_s["__s_code__"] = (
              df_s[ship_code_col]
              .astype(str)
              .str.strip()
              .str.upper()
              .str.replace(".0", "", regex=False)
          )
          df_c["__c_code__"] = (
              df_c[cust_code_col]
              .astype(str)
              .str.strip()
              .str.upper()
              .str.replace(".0", "", regex=False)
          )

          c_name_col = next(
              (
                  c
                  for c in df_c.columns
                  if "الاسم" in c or "name" in c.lower() or "اسم" in c
              ),
              None,
          )
          c_phone_col = next(
              (
                  c
                  for c in df_c.columns
                  if ("هاتف" in c or "عاتف" in c or "phone" in c.lower())
                  and "2" not in c
              ),
              None,
          )
          c_phone2_col = next(
              (
                  c
                  for c in df_c.columns
                  if ("هاتف" in c or "عاتف" in c or "phone" in c.lower())
                  and "2" in c
              ),
              None,
          )
          c_addr_col = next(
              (
                  c
                  for c in df_c.columns
                  if "عنوان" in c
                  or "address" in c.lower()
                  or "البض" in c
                  or "البظ" in c
              ),
              None,
          )
          c_city_col = next(
              (
                  c
                  for c in df_c.columns
                  if "مدينة" in c or "محافظ" in c or "city" in c.lower()
              ),
              None,
          )

          name_dict = (
              dict(zip(df_c["__c_code__"], df_c[c_name_col]))
              if c_name_col
              else {}
          )
          phone_dict = (
              dict(zip(df_c["__c_code__"], df_c[c_phone_col]))
              if c_phone_col
              else {}
          )
          phone2_dict = (
              dict(zip(df_c["__c_code__"], df_c[c_phone2_col]))
              if c_phone2_col
              else {}
          )
          addr_dict = (
              dict(zip(df_c["__c_code__"], df_c[c_addr_col]))
              if c_addr_col
              else {}
          )
          city_dict = (
              dict(zip(df_c["__c_code__"], df_c[c_city_col]))
              if c_city_col
              else {}
          )

          s_name_col = next(
              (
                  c
                  for c in df_s.columns
                  if "الاسم" in c and c != "__s_code__"
              ),
              "الاسم",
          )
          s_phone_col = next(
              (
                  c
                  for c in df_s.columns
                  if ("هاتف" in c or "عاتف" in c) and "2" not in c
              ),
              "رقم الهاتف",
          )
          s_phone2_col = next(
              (
                  c
                  for c in df_s.columns
                  if ("هاتف" in c or "عاتف" in c) and "2" in c
              ),
              "رقم الهاتف 2",
          )
          s_addr_col = next(
              (
                  c
                  for c in df_s.columns
                  if "عنوان" in c or "البض" in c or "البظ" in c
              ),
              "عنوان استلام البضاعة",
          )
          s_city_col = next(
              (c for c in df_s.columns if "مدينة" in c or "محافظ" in c),
              "المدينة",
          )

          if s_city_col not in df_s.columns:
            df_s[s_city_col] = "غير محدد"
          if s_addr_col not in df_s.columns:
            df_s[s_addr_col] = ""

          df_s[s_name_col] = (
              df_s["__s_code__"].map(name_dict).fillna(df_s.get(s_name_col))
          )
          df_s[s_phone_col] = (
              df_s["__s_code__"].map(phone_dict).fillna(df_s.get(s_phone_col))
          )
          if c_phone2_col:
            if s_phone2_col not in df_s.columns:
              df_s[s_phone2_col] = ""
            df_s[s_phone2_col] = (
                df_s["__s_code__"]
                .map(phone2_dict)
                .fillna(df_s.get(s_phone2_col))
            )

          mapped_addrs = df_s["__s_code__"].map(addr_dict)
          original_addrs = df_s.get(s_addr_col, pd.Series([""] * len(df_s)))
          df_s[s_addr_col] = mapped_addrs.combine_first(original_addrs).fillna(
              ""
          )

          df_s[s_city_col] = (
              df_s["__s_code__"]
              .map(city_dict)
              .fillna(df_s.get(s_city_col, "غير محدد"))
          )

          df_s.drop(columns=["__s_code__"], errors="ignore", inplace=True)
      except Exception:
        pass

    for col in list(df_s.columns):
      if col == "سعر الكيلو":
        df_s.rename(columns={col: "السعر"}, inplace=True)
        continue
      if col == "السعر":
        continue
      if "سعر" in str(col) and col != "السعر":
        df_s.rename(columns={col: "السعر"}, inplace=True)
        break

    return df_s


  st.markdown("---")
  st.header("🔍 فلتر الشحنات")
  selected_shipment_filter = "الكل"
  selected_code_filter = "الكل"
  selected_type_filter = "الكل"

  temp_df = load_and_merge_data(ship_mtime, cust_mtime)
  if temp_df is not None and not temp_df.empty:
    try:
      ship_col = next(
          (
              c
              for c in temp_df.columns
              if "شحنة" in str(c) or "shipment" in str(c).lower()
          ),
          temp_df.columns[0],
      )
      temp_df[ship_col] = (
          temp_df[ship_col]
          .fillna("بدون شحنة")
          .astype(str)
          .str.replace(".0", "", regex=False)
      )
      shipment_list = ["الكل"] + sorted(temp_df[ship_col].unique().tolist())
      selected_shipment_filter = st.selectbox(
          "اختر الشحنة للعرض:", shipment_list
      )

      filtered_temp_df = temp_df.copy()
      if selected_shipment_filter != "الكل":
        filtered_temp_df = filtered_temp_df[
            filtered_temp_df[ship_col] == selected_shipment_filter
        ]

      code_col = next(
          (
              c
              for c in filtered_temp_df.columns
              if "كود" in str(c) or "code" in str(c).lower()
          ),
          None,
      )
      if code_col:
        filtered_temp_df[code_col] = (
            filtered_temp_df[code_col]
            .fillna("بدون كود")
            .astype(str)
            .str.replace(".0", "", regex=False)
        )
        code_list = ["الكل"] + sorted(
            filtered_temp_df[code_col].unique().tolist()
        )
        selected_code_filter = st.selectbox(
            "اختر أو ابحث برقم الكود:", code_list
        )
        if selected_code_filter != "الكل":
          filtered_temp_df = filtered_temp_df[
              filtered_temp_df[code_col] == selected_code_filter
          ]

      type_col = next(
          (
              c
              for c in filtered_temp_df.columns
              if "نوع" in str(c) or "type" in str(c).lower()
          ),
          None,
      )
      if type_col:
        filtered_temp_df[type_col] = (
            filtered_temp_df[type_col]
            .fillna("غير محدد")
            .astype(str)
            .str.strip()
        )
        type_list = ["الكل"] + sorted(
            filtered_temp_df[type_col].unique().tolist()
        )
        selected_type_filter = st.selectbox("اختر نوع الشحنة:", type_list)
    except Exception:
      pass

active_data_file = shipment_path if os.path.exists(shipment_path) else None
active_template_file = template_path if os.path.exists(template_path) else None
active_logo = logo_path if os.path.exists(logo_path) else None

if active_data_file is not None and active_template_file is not None:
  try:
    df = load_and_merge_data(ship_mtime, cust_mtime)
    if df is None or df.empty:
      st.warning("⚠️ ملف البيانات فارغ أو لم يتم سحبه بنجاح من درايف.")
      st.stop()

    ship_col = next(
        (
            c
            for c in df.columns
            if "شحنة" in str(c) or "shipment" in str(c).lower()
        ),
        df.columns[0],
    )
    code_col = next(
        (c for c in df.columns if "كود" in str(c) or "code" in str(c).lower()),
        None,
    )
    type_col_name = next(
        (
            c
            for c in df.columns
            if "نوع" in str(c) or "type" in str(c).lower()
        ),
        None,
    )
    city_col_name = next(
        (
            c
            for c in df.columns
            if "مدينة" in str(c) or "محافظ" in str(c) or "city" in str(c).lower()
        ),
        "المدينة",
    )

    df[ship_col] = (
        df[ship_col]
        .fillna("بدون شحنة")
        .astype(str)
        .str.replace(".0", "", regex=False)
    )
    if code_col:
      df[code_col] = (
          df[code_col]
          .fillna("بدون كود")
          .astype(str)
          .str.replace(".0", "", regex=False)
      )
    if type_col_name:
      df[type_col_name] = (
          df[type_col_name].fillna("غير محدد").astype(str).str.strip()
      )
    if city_col_name in df.columns:
      df[city_col_name] = (
          df[city_col_name].fillna("غير محدد").astype(str).str.strip()
      )

    if selected_shipment_filter != "الكل":
      df = df[df[ship_col] == selected_shipment_filter]

    if selected_code_filter != "الكل" and code_col:
      df = df[df[code_col] == selected_code_filter]

    if selected_type_filter != "الكل" and type_col_name:
      df = df[df[type_col_name] == selected_type_filter]

    if df.empty:
      st.warning("⚠️ لا توجد بيانات مطابقة للفلاتر المحددة.")
      st.stop()

    today_date = datetime.date.today().strftime("%Y-%m-%d")

    logo_base64 = ""
    if active_logo and os.path.exists(active_logo):
      try:
        with open(active_logo, "rb") as img_file:
          logo_base64 = base64.b64encode(img_file.read()).decode("utf-8")
      except:
        pass

    name_col_for_clients = next(
        (c for c in df.columns if "الاسم" in c or "name" in c.lower()), None
    )
    if name_col_for_clients:
      total_clients_count = (
          df[name_col_for_clients]
          .dropna()
          .astype(str)
          .str.strip()
          .loc[lambda x: ~x.isin(["nan", "None", "", "عميل غير محدد"])]
          .nunique()
      )
      if total_clients_count == 0:
        total_clients_count = len(df)
    else:
      total_clients_count = len(df)

    weight_col = next(
        (c for c in df.columns if "وزن" in c or "weight" in c.lower()), None
    )
    total_weight_sum = (
        float(df[weight_col].sum())
        if weight_col and weight_col in df.columns
        else 0.0
    )

    cbm_col = next(
        (c for c in df.columns if "cbm" in c.lower() or "حجم" in c), None
    )
    total_cbm_sum = (
        float(df[cbm_col].sum()) if cbm_col and cbm_col in df.columns else 0.0
    )

    packages_col = next(
        (c for c in df.columns if "طرود" in c or "packages" in c.lower()), None
    )
    total_packages_count = (
        int(df[packages_col].sum())
        if packages_col and packages_col in df.columns
        else 0
    )

    price_col = next((c for c in df.columns if c == "السعر" or "سعر" in c), None)
    sales_col = next(
        (
            c
            for c in df.columns
            if "مبيعات" in c or "اجمالي" in c or "total" in c.lower()
        ),
        None,
    )

    if sales_col and sales_col in df.columns:
      total_sales_sum = float(df[sales_col].sum())
    elif price_col and weight_col:
      total_sales_sum = float((df[weight_col] * df[price_col]).sum())
    else:
      total_sales_sum = 0.0


    def generate_single_receipt_html(row_data):
      shipment = str(row_data.get(ship_col, "بدون شحنة")).strip()
      code = str(row_data.get(code_col, "بدون كود")).strip() if code_col else ""
      display_code = "" if code in ["بدون كود", "nan", "None"] else code
      display_shipment = (
          "" if shipment in ["بدون شحنة", "nan", "None"] else shipment
      )

      name = (
          str(row_data.get(name_col_for_clients, "عميل غير محدد")).strip()
          if name_col_for_clients
          else "عميل غير محدد"
      )
      if name in ["nan", "None", ""]:
        name = "عميل غير محدد"

      weight = (
          float(row_data.get(weight_col, 0) or 0) if weight_col else 0.0
      )
      cbm_value = (
          float(row_data.get(cbm_col, 0) or 0)
          if cbm_col and cbm_col in row_data
          else 0.0
      )
      packages = (
          int(float(row_data.get(packages_col, 0) or 0))
          if packages_col and packages_col in row_data
          else 0
      )
      price_per_kg = (
          float(row_data.get(price_col, 0) or 0) if price_col else 0.0
      )

      total_sales = (
          float(row_data.get(sales_col, 0) or 0)
          if sales_col and sales_col in row_data
          else 0.0
      )
      if total_sales == 0 and price_per_kg > 0 and weight > 0:
        total_sales = weight * price_per_kg

      phone_col = next(
          (
              c
              for c in df.columns
              if ("هاتف" in c or "عاتف" in c or "phone" in c.lower())
              and "2" not in c
          ),
          None,
      )
      phone = str(row_data.get(phone_col, "")).strip() if phone_col else ""
      if phone.endswith(".0"):
        phone = phone[:-2]
      phone = phone.replace("+", "").strip()
      if phone.startswith("964"):
        phone = phone[3:]
      formatted_phone = (
          f"+964 {phone}" if phone and phone not in ["nan", "None"] else ""
      )

      phone2_col = next(
          (
              c
              for c in df.columns
              if ("هاتف" in c or "عاتف" in c or "phone" in c.lower())
              and "2" in c
          ),
          None,
      )
      phone2 = str(row_data.get(phone2_col, "")).strip() if phone2_col else ""
      if phone2.endswith(".0"):
        phone2 = phone2[:-2]
      phone2 = phone2.replace("+", "").strip()
      if phone2.startswith("964"):
        phone2 = phone2[3:]
      formatted_phone2 = (
          f"+964 {phone2}" if phone2 and phone2 not in ["nan", "None"] else ""
      )

      combined_phones = formatted_phone
      if formatted_phone2:
        combined_phones = (
            f"{formatted_phone} / {formatted_phone2}"
            if formatted_phone
            else formatted_phone2
        )

      address_col = next(
          (
              c
              for c in df.columns
              if "عنوان" in c
              or "address" in c.lower()
              or "البض" in c
              or "البظ" in c
          ),
          None,
      )
      address = str(row_data.get(address_col, "")).strip() if address_col else ""
      if address in ["nan", "None"]:
        address = ""

      shipment_type = (
          str(row_data.get(type_col_name, "")).strip()
          if type_col_name
          else ""
      )
      if shipment_type in ["nan", "None"]:
        shipment_type = ""

      logo_img_tag = (
          f'<img src="data:image/png;base64,{logo_base64}" style="max-height:'
          ' 52px; max-width: 60px; margin-left: 10px; vertical-align:'
          ' middle;">'
          if logo_base64
          else ""
      )

      return f"""
            <div class="receipt-page" style="padding: 15px; font-family: 'Tahoma', Arial, sans-serif; direction: rtl; border: 2px solid #102a43; width: 100%; max-width: 148mm; margin: auto auto 20px auto; background: #ffffff; color: #102a43; box-sizing: border-box; page-break-after: always; break-after: page;">
                <table style="width: 100%; border-bottom: 2px solid #102a43; padding-bottom: 8px; margin-bottom: 12px; border-collapse: collapse;">
                    <tr>
                        <td style="text-align: right; vertical-align: middle;">
                            <div style="display: flex; align-items: center;">
                                {logo_img_tag}
                                <div>
                                    <h2 style="margin: 0; font-size: 15px; color: #102a43;">أطلس المحيط للتجارة العامة</h2>
                                    <p style="margin: 2px 0 0; font-size: 10px; color: #627d98;">OCEAN ATLAS GENERAL TRADING</p>
                                </div>
                            </div>
                        </td>
                        <td style="text-align: left; vertical-align: middle;">
                            <h3 style="margin: 0; font-size: 13px; color: #b45309;">وصل تسليم بضاعة</h3>
                            <p style="margin: 2px 0 0; font-size: 10px; color: #334e68;">Cargo Delivery Receipt</p>
                        </td>
                    </tr>
                </table>
                <table style="width: 100%; font-size: 11px; border-collapse: collapse; margin-bottom: 10px;">
                    <tr style="background-color: #f0f4f8;">
                        <td style="padding: 5px; border: 1px solid #bcccdc; width: 50%;"><strong>كود العميل:</strong> <span style="color: #b45309; font-weight: bold;">{display_code}</span></td>
                        <td style="padding: 5px; border: 1px solid #bcccdc; width: 50%;"><strong>رقم الشحنة:</strong> <span style="color: #b45309; font-weight: bold;">{display_shipment}</span></td>
                    </tr>
                    <tr>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>اسم العميل:</strong> <span style="font-weight: bold;">{name}</span></td>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>رقم الهاتف:</strong> <span style="direction: ltr; display: inline-block; font-weight: bold;">{combined_phones}</span></td>
                    </tr>
                    <tr style="background-color: #f0f4f8;">
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>عنوان الاستلام:</strong> <span style="color: #486581; font-weight: bold;">{address}</span></td>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>عدد الطرود:</strong> 📦 {packages} طرد</td>
                    </tr>
                    <tr>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>تاريخ الإصدار:</strong> <span style="color: #b45309; font-weight: bold;">{today_date}</span></td>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>الوزن الإجمالي:</strong> <span style="color: #102a43; font-weight: bold;">{weight} كغ</span></td>
                    </tr>
                    <tr style="background-color: #f0f4f8;">
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>نوع الشحنة:</strong> {shipment_type}</td>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>حجم الشحنة (CBM):</strong> <span style="color: #b45309; font-weight: bold;">{cbm_value:,.1f}</span></td>
                    </tr>
                    <tr>
                        <td style="padding: 5px; border: 1px solid #bcccdc;" colspan="2"><strong>السعر:</strong> {price_per_kg:,.1f} $</td>
                    </tr>
                    <tr style="background-color: #fef3c7;">
                        <td style="padding: 5px; border: 1px solid #f59e0b;" colspan="2"><strong>إجمالي المبيعات (الديون):</strong> <span style="color: #b45309; font-weight: bold; font-size: 12px;">{total_sales:,.1f} $</span> &nbsp;&nbsp;&nbsp;&nbsp;|&nbsp;&nbsp;&nbsp;&nbsp; <strong>طريقة الدفع:</strong> [ &nbsp; ] نقداً &nbsp;&nbsp; [ &nbsp; ] أجل</td>
                    </tr>
                </table>
                <div style="background-color: #fffbeb; border: 1px solid #fde68a; padding: 6px; border-radius: 4px; margin-bottom: 10px;">
                    <p style="margin: 0; font-size: 10px; color: #92400e; line-height: 1.3;"><strong>إقرار الاستلام:</strong><br>أقر أنا الموقع أدناه، بأنني استلمت البضاعة والشحنة المذكورة أعلاه كاملة، وبحالة سليمة وممتازة، ومطابقة لكافة الأوزان والأوصاف المدونة.</p>
                </div>
                <table style="width: 100%; font-size: 11px; margin-top: 5px; margin-bottom: 10px;">
                    <tr>
                        <td style="width: 50%; padding: 2px;"><strong>اسم المستلم:</strong><br><br>............................................</td>
                        <td style="width: 50%; padding: 2px; text-align: left;"><strong>توقيع وختم المستلم:</strong><br><br>............................................</td>
                    </tr>
                </table>
                <div style="border-top: 1px dashed #bcccdc; margin-top: 10px; padding-top: 6px; text-align: center; font-size: 9.5px; color: #334e68;">
                    <span>📍 العنوان: بغداد - المنصور - تقاطع الواد</span><span style="margin: 0 10px;">|</span><span style="direction: ltr; display: inline-block;">📞 هاتف: 07858588899 / 07814518989</span>
                </div>
            </div>
            """

    group_cols = [
        c
        for c in [
            ship_col,
            code_col,
            name_col_for_clients,
            type_col_name,
            city_col_name,
        ]
        if c and c in df.columns
    ]
    agg_dict = {}
    if weight_col and weight_col in df.columns:
      agg_dict[weight_col] = "sum"
    if cbm_col and cbm_col in df.columns:
      agg_dict[cbm_col] = "sum"
    if packages_col and packages_col in df.columns:
      agg_dict[packages_col] = "sum"
    if sales_col and sales_col in df.columns:
      agg_dict[sales_col] = "sum"

    for c in df.columns:
      if c not in group_cols and c not in agg_dict:
        agg_dict[c] = "first"

    if code_col and code_col in df.columns and len(group_cols) > 0:
      df_grouped = df.groupby(group_cols, as_index=False).agg(agg_dict)
    else:
      df_grouped = df.copy()

    if sales_col and sales_col in df_grouped.columns:
      df_grouped[sales_col] = df_grouped[sales_col].apply(
          lambda x: f"{float(x):,.1f}" if pd.notnull(x) else "0.0"
      )
    if price_col and price_col in df_grouped.columns:
      df_grouped[price_col] = df_grouped[price_col].apply(
          lambda x: f"{float(x):,.1f}" if pd.notnull(x) else "0.0"
      )

    st.markdown(
        """
        <style>
        .metric-card-1 { background-color: #eff6ff; border: 1px solid #bfdbfe; padding: 15px; border-radius: 10px; text-align: center; }
        .metric-card-2 { background-color: #f0fdf4; border: 1px solid #bbf7d0; padding: 15px; border-radius: 10px; text-align: center; }
        .metric-card-3 { background-color: #f5f3ff; border: 1px solid #ddd6fe; padding: 15px; border-radius: 10px; text-align: center; }
        .metric-card-4 { background-color: #fffbeb; border: 1px solid #fde68a; padding: 15px; border-radius: 10px; text-align: center; }
        .metric-card-5 { background-color: #fdf2f8; border: 1px solid #fbcfe8; padding: 15px; border-radius: 10px; text-align: center; }
        </style>
    """,
        unsafe_allow_html=True,
    )

    m1, m2, m3, m4, m5 = st.columns(5)
    with m1:
      st.markdown(
          f'<div class="metric-card-1"><p style="margin:0; color:#1e40af;'
          ' font-weight:bold; font-size:14px;">👥 عدد العملاء</p><h3'
          f' style="margin:5px 0 0; color:#1e3a8a;'
          f' font-size:20px;">{total_clients_count} عميل</h3></div>',
          unsafe_allow_html=True,
      )
    with m2:
      st.markdown(
          f'<div class="metric-card-2"><p style="margin:0; color:#166534;'
          ' font-weight:bold; font-size:14px;">📦 إجمالي الطرود</p><h3'
          f' style="margin:5px 0 0; color:#14532d;'
          f' font-size:20px;">{total_packages_count} طرد</h3></div>',
          unsafe_allow_html=True,
      )
    with m3:
      st.markdown(
          f'<div class="metric-card-3"><p style="margin:0; color:#5b21b6;'
          ' font-weight:bold; font-size:14px;">📐 إجمالي الحجم</p><h3'
          f' style="margin:5px 0 0; color:#4c1d95;'
          f' font-size:20px;">{total_cbm_sum:,.1f} CBM</h3></div>',
          unsafe_allow_html=True,
      )
    with m4:
      st.markdown(
          f'<div class="metric-card-4"><p style="margin:0; color:#92400e;'
          ' font-weight:bold; font-size:14px;">⚖️ الوزن الكلي</p><h3'
          f' style="margin:5px 0 0; color:#78350f;'
          f' font-size:20px;">{total_weight_sum:,.1f} كغ</h3></div>',
          unsafe_allow_html=True,
      )
    with m5:
      st.markdown(
          f'<div class="metric-card-5"><p style="margin:0; color:#9d174d;'
          ' font-weight:bold; font-size:14px;">💰 المبلغ الإجمالي'
          f' (الديون)</p><h3 style="margin:5px 0 0; color:#831843;'
          f' font-size:20px;">{total_sales_sum:,.1f} $</h3></div>',
          unsafe_allow_html=True,
      )

    st.markdown("<br>", unsafe_allow_html=True)

    st.subheader("📊 ملخص الإحصائيات والديون حسب المحافظات (المدن)")
    city_group_col = (
        city_col_name if city_col_name in df.columns else "المدينة"
    )
    if city_group_col not in df.columns:
      df["المدينة"] = "غير محدد"
      city_group_col = "المدينة"

    agg_city_dict = {}
    # حساب عدد العملاء الفريدين أو عدد الأكواد لكل مدينة
    if code_col and code_col in df.columns:
      agg_city_dict[code_col] = "nunique"
    elif name_col_for_clients and name_col_for_clients in df.columns:
      agg_city_dict[name_col_for_clients] = "nunique"
    else:
      agg_city_dict[city_group_col] = "count"

    if packages_col and packages_col in df.columns:
      agg_city_dict[packages_col] = "sum"
    if cbm_col and cbm_col in df.columns:
      agg_city_dict[cbm_col] = "sum"
    if sales_col and sales_col in df.columns:
      agg_city_dict[sales_col] = "sum"
    elif weight_col and price_col:
      df["__calc_sales__"] = df[weight_col] * df[price_col]
      agg_city_dict["__calc_sales__"] = "sum"

    if agg_city_dict:
      df_city_summary = df.groupby(city_group_col, as_index=False).agg(
          agg_city_dict
      )
      
      # إعادة تسمية الأعمدة بشكل منظم مع إضافة حقل عدد العملاء
      rename_mapping = {}
      if code_col in df_city_summary.columns:
        rename_mapping[code_col] = "عدد العملاء"
      elif name_col_for_clients in df_city_summary.columns:
        rename_mapping[name_col_for_clients] = "عدد العملاء"
      else:
        rename_mapping[city_group_col] = "عدد العملاء"

      if "__calc_sales__" in df_city_summary.columns:
        rename_mapping["__calc_sales__"] = "إجمالي الديون / المبيعات ($)"
      if sales_col and sales_col in df_city_summary.columns:
        rename_mapping[sales_col] = "إجمالي الديون / المبيعات ($)"
      if packages_col and packages_col in df_city_summary.columns:
        rename_mapping[packages_col] = "إجمالي الطرود"
      if cbm_col and cbm_col in df_city_summary.columns:
        rename_mapping[cbm_col] = "إجمالي الحجم (CBM)"

      df_city_summary.rename(columns=rename_mapping, inplace=True)

      if "إجمالي الديون / المبيعات ($)" in df_city_summary.columns:
        df_city_summary["إجمالي الديون / المبيعات ($)"] = df_city_summary[
            "إجمالي الديون / المبيعات ($)"
        ].apply(lambda x: f"{float(x):,.1f}" if pd.notnull(x) else "0.0")
      if "إجمالي الحجم (CBM)" in df_city_summary.columns:
        df_city_summary["إجمالي الحجم (CBM)"] = df_city_summary[
            "إجمالي الحجم (CBM)"
        ].apply(lambda x: f"{float(x):,.1f}" if pd.notnull(x) else "0.0")

      df_city_summary.insert(0, "التسلسل", range(1, len(df_city_summary) + 1))
      city_table_html = df_city_summary.to_html(
          classes="custom-table", index=False, escape=False
      )
      st.html(f"""<div class="custom-table-container">{city_table_html}</div>""")
    else:
      city_table_html = "<p>لا توجد بيانات كافية لعرض ملخص المحافظات.</p>"

    st.markdown("<br>", unsafe_allow_html=True)
    st.subheader(
        f"📋 جدول تفاصيل الشحنة المعروضة: [{selected_shipment_filter}] - النوع:"
        f" [{selected_type_filter}]"
    )

    display_table_df = df_grouped.copy()
    display_table_df.insert(0, "التسلسل", range(1, len(display_table_df) + 1))
    table_html = display_table_df.to_html(
        classes="custom-table", index=False, escape=False
    )

    custom_table_styling = f"""
    <style>
        .custom-table-container {{
            max-height: 450px;
            overflow-x: auto;
            overflow-y: auto;
            border: 1px solid #bcccdc;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            margin-bottom: 20px;
        }}
        .custom-table {{
            width: 100%;
            border-collapse: collapse;
            font-family: 'Tahoma', Arial, sans-serif;
            font-size: 13px;
            direction: rtl;
            background-color: #ffffff;
            color: #102a43;
        }}
        .custom-table th {{
            background-color: #102a43 !important;
            color: #ffffff !important;
            text-align: right;
            padding: 12px 10px;
            font-weight: bold;
            border-bottom: 2px solid #0b1e33;
            position: sticky;
            top: 0;
            z-index: 10;
            white-space: nowrap;
        }}
        .custom-table td {{
            padding: 10px 10px;
            border-bottom: 1px solid #e2e8f0;
            text-align: right;
        }}
        .custom-table tr:nth-child(even) {{
            background-color: #f8fafc;
        }}
        .custom-table tr:hover {{
            background-color: #f1f5f9;
        }}
    </style>
    <div class="custom-table-container">
        {table_html}
    </div>
    """
    st.html(custom_table_styling)

    table_pdf_html_component = f"""
        <!DOCTYPE html>
        <html lang="ar" dir="rtl">
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{ size: A4 landscape; margin: 10mm; }}
                body {{ font-family: 'Tahoma', Arial, sans-serif; direction: rtl; color: #102a43; margin: 0; padding: 0; background: transparent; }}
                .export-btn {{
                    background-color: #102a43;
                    color: white;
                    padding: 12px 24px;
                    border: none;
                    border-radius: 6px;
                    cursor: pointer;
                    font-weight: bold;
                    font-size: 14px;
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    font-family: 'Tahoma', Arial, sans-serif;
                }}
                .export-btn:hover {{
                    background-color: #0b1e33;
                }}
            </style>
        </head>
        <body>
            <button class="export-btn" onclick="exportTablePDF()">📄 تصدير الجدول الحالي وتقرير المحافظات إلى PDF</button>
            <script>
                const tableContent = `{table_html.replace('`', '\\`').replace('$', '\\$')}`;
                const citySummaryContent = `{city_table_html.replace('`', '\\`').replace('$', '\\$')}`;
                const filterInfo = 'الشحنة: {selected_shipment_filter} | النوع: {selected_type_filter}';
                const totalClients = '{total_clients_count} عميل';
                const totalPackages = '{total_packages_count} طرد';
                const totalCbm = '{total_cbm_sum:,.1f} CBM';
                const totalWeight = '{total_weight_sum:,.1f} كغ';
                const totalSales = '{total_sales_sum:,.1f} $';

                function exportTablePDF() {{
                    var w = window.open('', '', 'height=900,width=1200');
                    w.document.write(`
                        <!DOCTYPE html>
                        <html lang="ar" dir="rtl">
                        <head>
                            <meta charset="UTF-8">
                            <title>تقرير جدول الشحنات والمحافظات - أطلس</title>
                            <style>
                                @page {{ size: A4 landscape; margin: 5mm; }}
                                body {{ font-family: Tahoma, Arial, sans-serif; direction: rtl; color: #102a43; padding: 5px; }}
                                h3 {{ color: #102a43; margin-top: 10px; font-size: 13px; border-bottom: 2px solid #102a43; padding-bottom: 3px; }}
                                .metrics-grid {{ display: flex; justify-content: space-between; gap: 5mm; margin-bottom: 10px; }}
                                .metric-box {{ flex: 1; padding: 6px; border-radius: 6px; text-align: center; border: 1px solid #cbd5e1; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
                                .box-1 {{ background-color: #eff6ff !important; border-color: #bfdbfe; color: #1e40af; }}
                                .box-2 {{ background-color: #f0fdf4 !important; border-color: #bbf7d0; color: #166534; }}
                                .box-3 {{ background-color: #f5f3ff !important; border-color: #ddd6fe; color: #5b21b6; }}
                                .box-4 {{ background-color: #fffbeb !important; border-color: #fde68a; color: #92400e; }}
                                .box-5 {{ background-color: #fdf2f8 !important; border-color: #fbcfe8; color: #9d174d; }}
                                .metric-title {{ font-size: 10px; font-weight: bold; margin-bottom: 2px; }}
                                .metric-val {{ font-size: 12px; font-weight: bold; margin: 0; }}
                                table {{ width: 100% !important; border-collapse: collapse; font-size: 8.5px !important; margin-top: 5px; table-layout: fixed; }}
                                th, td {{ padding: 4px 3px !important; border: 1px solid #cbd5e1; text-align: right; overflow: hidden; word-wrap: break-word; }}
                                th {{ background-color: #102a43 !important; color: #ffffff !important; font-weight: bold; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
                                tr:nth-child(even) {{ background-color: #f8fafc; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
                            </style>
                        </head>
                        <body>
                            <div class="metrics-grid">
                                <div class="metric-box box-1"><div class="metric-title">👥 عدد العملاء</div><div class="metric-val">${{totalClients}}</div></div>
                                <div class="metric-box box-2"><div class="metric-title">📦 إجمالي الطرود</div><div class="metric-val">${{totalPackages}}</div></div>
                                <div class="metric-box box-3"><div class="metric-title">📐 إجمالي الحجم</div><div class="metric-val">${{totalCbm}}</div></div>
                                <div class="metric-box box-4"><div class="metric-title">⚖️ الوزن الكلي</div><div class="metric-val">${{totalWeight}}</div></div>
                                <div class="metric-box box-5"><div class="metric-title">💰 المبلغ الإجمالي</div><div class="metric-val">${{totalSales}}</div></div>
                            </div>
                            <h3>📊 ملخص الإحصائيات حسب المحافظات</h3>
                            ${{citySummaryContent}}
                            <h3 style="margin-top: 15px;">📋 تفاصيل الشحنات (${{filterInfo}})</h3>
                            ${{tableContent}}
                        </body>
                        </html>
                    `);
                    w.document.close();
                    w.focus();
                    setTimeout(() => {{ w.print(); w.close(); }}, 600);
                }}
            </script>
        </body>
        </html>
    """
    st.components.v1.html(table_pdf_html_component, height=55)
    st.markdown("---")

    # --- تحضير كود الوصولات لطباعتها دفعة واحدة عبر مكون HTML مستقل يضمن تفاعل الزر تماماً ---
    all_html_batch = ""
    for _, row in df.iterrows():
      all_html_batch += generate_single_receipt_html(row)

    batch_print_component = f"""
        <!DOCTYPE html>
        <html lang="ar" dir="rtl">
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{ size: A5; margin: 5mm; }}
                body {{ font-family: 'Tahoma', Arial, sans-serif; direction: rtl; margin: 0; padding: 0; background: transparent; }}
                .batch-btn {{
                    background-color: #b45309;
                    color: white;
                    padding: 12px 24px;
                    border: none;
                    border-radius: 6px;
                    cursor: pointer;
                    font-weight: bold;
                    font-size: 14px;
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    font-family: 'Tahoma', Arial, sans-serif;
                }}
                .batch-btn:hover {{
                    background-color: #92400e;
                }}
            </style>
        </head>
        <body>
            <button class="batch-btn" onclick="printAllBatch()">🖨️ طباعة الوصولات المعروضة دفعة واحدة (مقاس A5)</button>
            <script>
                const masterContent = `{all_html_batch.replace('`', '\\`').replace('$', '\\$')}`;
                function printAllBatch() {{
                    var w = window.open('', '', 'height=900,width=800');
                    w.document.write('<html><head><style>@page {{ size: A5; margin: 5mm; }} body {{ direction: rtl; font-family: Tahoma; }}</style></head><body>' + masterContent + '</body></html>');
                    w.document.close();
                    w.focus();
                    setTimeout(() => {{ w.print(); w.close(); }}, 600);
                }}
            </script>
        </body>
        </html>
    """
    st.components.v1.html(batch_print_component, height=55)
    st.markdown("---")

    for index, row in df.iterrows():
      shipment = str(row.get(ship_col, "بدون شحنة")).strip()
      code = str(row.get(code_col, "بدون كود")).strip() if code_col else ""
      display_code = "" if code in ["بدون كود", "nan", "None"] else code
      display_shipment = (
          "" if shipment in ["بدون شحنة", "nan", "None"] else shipment
      )
      name = (
          str(row.get(name_col_for_clients, "عميل غير محدد")).strip()
          if name_col_for_clients
          else "عميل غير محدد"
      )
      if name in ["nan", "None", ""]:
        name = "عميل غير محدد"

      weight = (
          float(row.get(weight_col, 0) or 0) if weight_col else 0.0
      )
      price_per_kg = (
          float(row.get(price_col, 0) or 0) if price_col else 0.0
      )
      sales_col_val = (
          float(row.get(sales_col, 0) or 0)
          if sales_col and sales_col in row
          else 0.0
      )
      if sales_col_val == 0 and price_per_kg > 0 and weight > 0:
        sales_col_val = weight * price_per_kg

      with st.expander(
          f"📄 وصل العميل: {name} | كود: {display_code or 'بدون'} | الشحنة:"
          f" {display_shipment} | الإجمالي: {sales_col_val:,.1f} $"
      ):
        try:
          wb = openpyxl.load_workbook(active_template_file)
          ws = wb.active
          ws["B4"] = display_code
          ws["D4"] = today_date
          ws["B5"] = name
          ws["B6"] = (
              str(
                  row.get(
                      next(
                          (
                              c
                              for c in df.columns
                              if "عنوان" in c or "البض" in c
                          ),
                          "",
                      ),
                      "",
                  )
              ).strip()
              if next(
                  (c for c in df.columns if "عنوان" in c or "البض" in c), None
              )
              else ""
          )

          p_col = next(
              (
                  c
                  for c in df.columns
                  if ("هاتف" in c or "عاتف" in c) and "2" not in c
              ),
              None,
          )
          p_val = str(row.get(p_col, "")).strip() if p_col else ""
          if p_val.endswith(".0"):
            p_val = p_val[:-2]
          p_val = p_val.replace("+", "").strip()
          if p_val.startswith("964"):
            p_val = p_val[3:]

          ws["D5"] = f"+964 {p_val}" if p_val else ""
          ws["B7"] = display_shipment
          ws["D6"] = int(
              float(
                  row.get(
                      next(
                          (c for c in df.columns if "طرود" in c), None
                      ),
                      0,
                  )
                  or 0
              )
          )
          ws["B8"] = (
              str(row.get(type_col_name, "")).strip()
              if type_col_name
              else ""
          )
          ws["D7"] = weight

          output = io.BytesIO()
          wb.save(output)
          output.seek(0)
        except:
          output = io.BytesIO()

        single_html = generate_single_receipt_html(row)
        file_name_id = f"Shipment_{display_shipment}_Client_{name}".replace(
            " ", "_"
        )

        st.download_button(
            label="📥 تنزيل إكسل الوصل",
            data=output,
            file_name=f"Delivery_Receipt_{file_name_id}.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            key=f"dl_{index}",
        )
        st.markdown("<br>", unsafe_allow_html=True)
        st.components.v1.html(
            f"""<div style="direction:rtl">{single_html}</div><button style="background:#102a43;color:white;padding:12px 20px;border:none;border-radius:6px;cursor:pointer;font-weight:bold;margin-top:15px;" onclick="window.print()">🖨️ طباعة هذا الوصل</button>""",
            height=700,
            scrolling=True,
        )
      st.markdown("---")

  except Exception as e:
    st.error(f"حدث خطأ أثناء معالجة الملفات: {e}")
else:
  st.info(
      "الرجاء التأكد من صلاحية الوصول للملفات والضغط على زر (تحديث البيانات و"
      "سحبها من درايف)."
  )
